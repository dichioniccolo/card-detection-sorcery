import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from PIL import Image

from geometry import locate_cells
from card_mask import extract_card
from label_ocr import read_label
from deposit_metrics import analyze_card, DEFAULT_DPI

# Volume di applicazione (L/ha) usato nelle prove di riferimento.
DEFAULT_DROP_SIZE = 200

OUTPUT_FIELDS = [
    "HEIGHT",
    "PLANT",
    "REPLICA",
    "SIDE",
    "DIRECTION",
    "TEST",
    "DROP SIZE",
    "DV01",
    "DV05",
    "DV09",
    "COVERAGE",
    "IMAGE AREA",
    "TOTAL DEPOSIT",
    "DROP DENSITY",
    "µL",
    "quality_flag",
    "largest_component_frac",
    "background_floor_gray",
    "label_ok",
    "label_raw_text",
    "source_file",
    "card_index",
]


def process_sheet(image_path: str, dpi: float = DEFAULT_DPI, drop_size=DEFAULT_DROP_SIZE,
                  relaxed: bool = False):
    """Ritorna (righe, note). `note` elenca i recuperi fatti in modalita'
    permissiva, da riportare all'utente: sono card da controllare a mano."""
    image = Image.open(image_path)
    cells, grid_note = locate_cells(image_path, relaxed=relaxed)
    notes = [grid_note] if grid_note else []

    rows = []
    for cell in cells:
        try:
            row = _process_cell(image, cell, dpi, drop_size, image_path)
        except Exception as e:
            if not relaxed:
                raise
            # In modalita' permissiva una card illeggibile non fa perdere le
            # altre tre: esce una riga senza metriche, marcata nel quality_flag.
            notes.append(f"card {cell.row_index + 1} non elaborata: {e}")
            rows.append(_empty_row(cell, drop_size, image_path, str(e)))
            continue
        rows.append(row)

    return rows, notes


def _empty_row(cell, drop_size, image_path, err):
    row = {f: "" for f in OUTPUT_FIELDS}
    row["DROP SIZE"] = drop_size
    row["quality_flag"] = f"CARD_NON_ELABORATA: {err}"
    row["label_ok"] = False
    row["source_file"] = image_path
    row["card_index"] = cell.row_index + 1
    return row


def _process_cell(image, cell, dpi, drop_size, image_path):
    card_crop = image.crop(cell.card_box).convert("RGB")
    card_rgb = np.array(card_crop)
    card_mask, _box = extract_card(card_rgb)

    label = read_label(image, cell.label_box)

    metrics = analyze_card(card_rgb, card_mask, dpi=dpi)

    row = {
        "HEIGHT": label.get("height"),
        "PLANT": label.get("plant"),
        "REPLICA": label.get("replica"),
        "SIDE": label.get("side"),
        "DIRECTION": label.get("direction"),
        "TEST": label.get("test"),
        # Volume di applicazione del trattamento (L/ha): metadato
        # sperimentale, non ricavabile dall'immagine.
        "DROP SIZE": drop_size,
        "DV01": round(metrics["dv01_um"], 1),
        "DV05": round(metrics["dv05_um"], 1),
        "DV09": round(metrics["dv09_um"], 1),
        "COVERAGE": round(metrics["coverage_pct"], 2),
        "IMAGE AREA": round(metrics["image_area_cm2"], 2),
        "TOTAL DEPOSIT": metrics["total_deposit_counted"],
        "DROP DENSITY": round(metrics["deposits_per_cm2"], 1),
        "µL": round(metrics["ul_cm2"], 3),
        "quality_flag": metrics["quality_flag"],
        # diagnostici del quality_flag: servono a capire perche' e' scattato
        "largest_component_frac": round(metrics["largest_component_frac"], 4),
        "background_floor_gray": round(metrics["background_floor_gray"], 1),
        "label_ok": label["ok"],
        "label_raw_text": label["raw_text"],
        "source_file": image_path,
        "card_index": cell.row_index + 1,
    }
    return row


# Campi che identificano una card: due righe con la stessa terna
# etichetta/lato/verso/tesi sono lo stesso rilievo estratto due volte.
DUP_KEY_FIELDS = ["HEIGHT", "PLANT", "REPLICA", "SIDE", "DIRECTION", "TEST"]


def find_duplicates(rows: list):
    """Righe con la stessa etichetta all'interno dello stesso output.

    Ritorna [(chiave, [righe]), ...] per le sole chiavi con piu' di una riga.
    Le righe senza etichetta letta (label_ok falso o campi vuoti) non entrano:
    non sono duplicati, sono etichette da rileggere a mano.
    """
    groups = {}
    for row in rows:
        values = [row.get(f) for f in DUP_KEY_FIELDS]
        if not row.get("label_ok") or any(v in (None, "") for v in values):
            continue
        groups.setdefault(tuple(values), []).append(row)
    return [(k, v) for k, v in groups.items() if len(v) > 1]


def format_duplicates(dups: list) -> list:
    """Righe di testo pronte per il log/console."""
    lines = []
    for key, rows in dups:
        label = " ".join(str(v) for v in key)
        lines.append(f"  {label} ({len(rows)} volte):")
        for r in rows:
            lines.append(f"      {r['source_file']} card {r['card_index']}")
    return lines


def write_xlsx(rows: list, out_path: str):
    """Una riga per card, colonne in ordine `OUTPUT_FIELDS`."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DepositScan"
    ws.append(OUTPUT_FIELDS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append([row.get(f, "") for f in OUTPUT_FIELDS])
    for i, name in enumerate(OUTPUT_FIELDS, start=1):
        width = max(len(name) + 2, 12)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    wb.save(out_path)
